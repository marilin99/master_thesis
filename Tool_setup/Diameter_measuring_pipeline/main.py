#!/usr/bin/python3
# CODE for checking a specific folder and running the diameter measuring pipeline #
# libraries
import cv2
import numpy as np 
import time 
import os
import re

## helper functions ##
from classical_segmenter import classical_segment
from scale_obtain import scale_obtain 
from thinner import thinner, thinner_2k_5k
from point_picker import point_picker
from dm_finder import dm_finder

## adding data to the xlsx file ##
from openpyxl import load_workbook
from datetime import datetime

# first_part = "/run/user/1000/gvfs/smb-share:server=gaia.domenis.ut.ee,share=mvfa/"
# ## To be processed SEMs folder on the Network Drive
# # ORIG_PATH = r"G:/Automaatika/SEM_input/"

# # ## moving files 
# # SRC_PATH = r"G:\Automaatika\SEM_input"
# # TARGET_PATH = r"G:\Automaatika\SEM_processed"
# # txt_TARGET_PATH = r"G:/Automaatika/SEM_processed/"

# # ## Excel workbook path ## 
# # filename = r"G:\Automaatika\SEM_analysis.xlsx"

# ORIG_PATH = first_part + "/Automaatika/SEM_input/"

# ## moving files 
# SRC_PATH = first_part + r"Automaatika\SEM_input"
# TARGET_PATH = first_part + r"Automaatika\SEM_processed"
# txt_TARGET_PATH = first_part + r"Automaatika/SEM_processed/"

# ## Excel workbook path ## 
# filename = first_part + "Automaatika\SEM_analysis.xlsx"

ORIG_PATH = "/home/marilin/Documents/ESP/data/automaatika_test/"
TARGET_PATH =  "/home/marilin/Documents/ESP/data/automaatika_test_output/"


FILES = os.listdir(ORIG_PATH) 
#print(FILES)

for f in FILES:

    if f.endswith("png") or f.endswith("jpg") or f.endswith("tif"):

        # dist values in pixels 
        dist_bool = "No" 
        print("file name", f)

        core_name = f[:-4]

        start_time = time.time()
        PATH_1 = ORIG_PATH+f
        # assuming an extension is 4 characters (including the ".")

        core_name = f[:-4]


        ## CLASSICAL segmentation ##
        # otsu thresholding for 2k and 5k and 500x if specified in path
        if re.search("(5k|2k|500x|500k)", PATH_1) != None:
            
            segmented_im = cv2.threshold(cv2.imread(PATH_1, 0), 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1]
            
            dist, thinned = thinner_2k_5k(segmented_im)
        
        else:
            # classical segmentation for other magnifications
            segmented_im = classical_segment(PATH_1)

            dist, thinned = thinner(segmented_im)
        
        cv2.imwrite(f"{TARGET_PATH}{core_name}_segmented.png", np.uint8(segmented_im))
            
        ### END OF CLASSICAL segmentation ###
        
        # 100 measurements per image 
        pt_s = point_picker(segmented_im, 100)
        
        # val, unit, px amount - from original image
        scales = scale_obtain(PATH_1)
        try:
            if len(scales) == 3: 
                if scales[1] == "um":
                    nano_per_px = int(scales[0]) * 1000 / int(scales[-1])
                # scale = nm 
                elif scales[1] == "nm": 
                    nano_per_px = int(scales[0]) / int(scales[-1])

            else:
                dist_bool = "Yes"
                nano_per_px = 1

        except Exception: 
            # adding a file for scale exception
            nano_per_px = 1
            dist_bool = "Yes"
            #with open(f"{txt_TARGET_PATH}{core_name}.txt", "w+") as file_ex:
            #    file_ex.write("Scaling off, check the image, adding distance in pixels for now")
            #continue


        # leaving the lower part (scale included) in for now
        h,w = segmented_im.shape[:2]

        first_dm_s, first_excs, coords  = dm_finder(thinned, dist, segmented_im, pt_s, h,w,nano_per_px) #[:3]

        ## TMP VISUALIZATION ##

        rgb_im = cv2.imread(PATH_1)

        start_color = (0, 0, 0)
        end_color = (255,0,0)  

        for combo in coords:
            #combo = eval(combo)

            start_point = (combo[1], combo[0])
            mid_point = (combo[3], combo[2])

            # testing the actual end point
            end_point_x = (2 * mid_point[0]) - start_point[0]
            end_point_y = (2 * mid_point[1]) - start_point[1]

            end_point = (end_point_x, end_point_y)

            # px dist
            distance1 = np.sqrt((end_point[1] - start_point[1]) ** 2 + (end_point[0] - start_point[0]) ** 2)

            cv2.line(rgb_im, start_point, mid_point, start_color, 2)
            cv2.line(rgb_im, mid_point, end_point, (0,255,0), 2)

            cv2.circle(rgb_im, mid_point, 3, end_color, -1)


        image = cv2.cvtColor(rgb_im, cv2.COLOR_RGB2BGR)

        cv2.imwrite(f"{TARGET_PATH}{core_name}_w_segments.png", np.uint8(image))
        ###

        ### saving values to the XLSX file ###
        # sheet_name = datetime.now().strftime("%B") +" "+datetime.now().strftime("%Y")

        # new_row = [core_name, str(first_dm_s), np.mean(first_dm_s), np.std(first_dm_s), np.median(first_dm_s), time.time() - start_time, dist_bool]

        # wb = load_workbook(filename)

        # if sheet_name in wb.sheetnames:
        #     ws = wb.worksheets[-1] # select last worksheet
        # else:
        #     wb.create_sheet(sheet_name)
        #     headers_row =  ["File path", "Diameter measures (nm)", "Mean (nm)", "Standard deviation (nm)", "Median (nm)", "Runtime (s)", "Values in pixels"]
        #     ws = wb.worksheets[-1]
        #     ws.append(headers_row)

        # ws.append(new_row)
        # wb.save(filename)
        # wb.close()

        # start_time = time.time()

        # ## moving the file to the target path ## 

        # os.rename(os.path.join(SRC_PATH, f), os.path.join(TARGET_PATH, f))


    #####
    

