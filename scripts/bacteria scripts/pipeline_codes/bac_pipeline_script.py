import os.path
import numpy as np
import cv2
import matplotlib.pyplot as plt
from scipy import ndimage
from scipy.ndimage import *
from skimage.morphology import skeletonize, medial_axis, disk
from skimage.feature import peak_local_max
import time 
from varname import argname
from natsort import natsorted

### helper functions ###
from bac_threshold import works_for_both
from peak_detection import peaker 
from czi_pipe_convert import converter

## adding data to the xlsx file ##
from openpyxl import load_workbook
from datetime import datetime


### variables ###
base_im_r, base_im_g = np.zeros((512,512), dtype=np.uint8), np.zeros((512,512), dtype=np.uint8)
start_time = time.time()

# To be processed CZIs folder on the Network Drive
#PATH = "/home/marilin/Documents/ESP/data/SYTO_PI_conversion/"
PATH = "/home/marilin/Documents/ESP/data/bacteria_tests/test_pipeline/"

# Processed CZI folder on the Network Drive 
TARGET_PATH = "TODO"

## Excel workbook path ## 
filename = "/home/marilin/Documents/ESP/data/bacteria_tests/bacteria_demo.xlsx"

FILES = os.listdir(PATH) 

bac_counting = {}


## check what is in the folder
start_time = time.time()


# filtering out czi and not control values 
for f in natsorted(FILES):

    if f.endswith(".czi") and "control" not in f:
        unique_file = f.split(".czi")[0]

        # empty file for every czi file 
        base_im_r, base_im_g = np.zeros((512,512), dtype=np.uint8), np.zeros((512,512), dtype=np.uint8)

        file = PATH+f
        # czi converter returns dict.items() 

        for color, stacks in converter(file):

            if "red" in color:

                # iterating through the stacks
                for im in stacks:
                    
                    shape_red = im.shape
                    red_chan = cv2.resize(im, (512,512))
                    
                    # centrosymmetric structure
                    labeled_array, num_features_r = label(peaker(*works_for_both(red_chan))[0])

                    # adding peaks to the empty image
                    base_im_r += peaker(*works_for_both(red_chan))[1]

            elif "green" in color:
                # iterating through the stacks

                for im in stacks:
                    shape_green = im.shape
                    green_chan = cv2.resize(im, (512,512))
                    
                    # centrosymmetric structure
                    labeled_array, num_features_g = label(peaker(*works_for_both(green_chan))[0])

                    base_im_g += peaker(*works_for_both(green_chan))[1]
        
        # full amount of bacteria over all the stacks 
        bac_counting[f"{unique_file}_red"] = label(base_im_r)[1]
        bac_counting[f"{unique_file}_green"] = label(base_im_g)[1]


  

        ### saving values to the XLSX file ###
        # current month + year
        sheet_name = datetime.now().strftime("%B") +" "+datetime.now().strftime("%Y")

        # type bac not possible right now 
        nr_of_stacks = len(stacks)
        green_red = [bac_counting[f"{unique_file}_green"], bac_counting[f"{unique_file}_red"]]
        green_red_ratio =  label(base_im_g)[1] / label(base_im_r)[1]
        decision = np.where(green_red_ratio<1, "Dead", "Alive").item()
        runtime =  time.time() - start_time

        new_row = [f, nr_of_stacks, str(green_red), decision,runtime]

        wb = load_workbook(filename)

        if sheet_name in wb.sheetnames:

            ws = wb.worksheets[-1] # select last worksheet
        else:
            wb.create_sheet(sheet_name)
            headers_row =  ["File path", "Nr of stacks", "Amount of green, red bacteria", "Dead/alive", "Runtime (s)"]
            ws = wb.worksheets[-1]
            ws.append(headers_row)

        ws.append(new_row)
        wb.save(filename)
        wb.close()

    start_time = time.time()

    ### moving the CZI file to the processed folder

    # TODO 