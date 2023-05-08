import os.path
import numpy as np
import cv2
import matplotlib.pyplot as plt
from scipy import ndimage
from scipy.ndimage import *
from skimage.morphology import skeletonize, medial_axis, disk
from skimage.feature import peak_local_max
import time 
from varname import argname
from natsort import natsorted

### helper functions ###
from bac_threshold import works_for_both
from peak_detection import peaker 
from czi_pipe_convert import converter

## adding data to the xlsx file ##
from openpyxl import load_workbook
from datetime import datetime


### variables ###
base_im_r, base_im_g = np.zeros((512,512), dtype=np.uint8), np.zeros((512,512), dtype=np.uint8)
start_time = time.time()

# To be processed CZIs folder on the Network Drive
#PATH = "/home/marilin/Documents/ESP/data/FM_NO_SYTO/"
PATH = "/home/marilin/Documents/ESP/data/SYTO_PI/"

# Processed CZI folder on the Network Drive 
TARGET_PATH = "/home/marilin/Documents/ESP/data/bacteria_tests/test_pipeline_res/"

## Excel workbook path ## 
filename = "/home/marilin/Documents/ESP/data/bacteria_tests/bacteria_demo.xlsx"

FILES = os.listdir(PATH) 

bac_counting = {}
intensity_dict = {}

## check what is in the folder
start_time = time.time()


# filtering out czi and not control values 
for f in natsorted(FILES):

    if f.endswith(".czi") and "control" not in f:
        unique_file = f.split(".czi")[0]

        # empty file for every czi file 
        base_im_r, base_im_g = np.zeros((512,512), dtype=np.uint8), np.zeros((512,512), dtype=np.uint8)
        tmp_im_r, tmp_im_g = np.zeros((512,512), dtype=np.uint8), np.zeros((512,512), dtype=np.uint8)

        file = PATH+f
        # czi converter returns dict.items() 

        for color, stacks in converter(file):
            if "red" in color:

                # iterating through the stacks
                for i, im in enumerate(stacks):
                    #tmp_im_r = np.zeros((512,512), dtype=np.uint8)

                    shape_red = im.shape
                    red_chan = cv2.resize(im, (512,512))
                    
                    # centrosymmetric structure
                    labeled_array, num_features_r = label(peaker(*works_for_both(red_chan))[0])

                    # adding peaks to the empty image
                    base_im_r += peaker(*works_for_both(red_chan))[1]

                    tmp_im_r += peaker(*works_for_both(red_chan))[2]

            elif "green" in color:
                # iterating through the stacks

                for j, im in enumerate(stacks):

                    shape_green = im.shape
                    green_chan = cv2.resize(im, (512,512))
                    
                    # centrosymmetric structure
                    labeled_array, num_features_g = label(peaker(*works_for_both(green_chan))[0])

                    base_im_g += peaker(*works_for_both(green_chan))[1]

                    tmp_im_g += peaker(*works_for_both(green_chan))[2]
                  
                
        
        # cv2.imwrite(f"{TARGET_PATH}{unique_file}_th_r.png", tmp_im_r)
        # cv2.imwrite(f"{TARGET_PATH}{unique_file}_th_g.png", tmp_im_g)
        
        red_c = tuple(zip(*np.where(tmp_im_r>0)))
        green_c = tuple(zip(*np.where(tmp_im_g>0)))

        # full amount of bacteria over all the stacks 
        bac_counting[f"{unique_file}_red"] = label(base_im_r)[1]
        bac_counting[f"{unique_file}_green"] = label(base_im_g)[1]

        labeled_array_r, num_features_r = label(base_im_r)
        labeled_array_g, num_features_g = label(base_im_g)

        uniq_vals_r = np.unique(labeled_array_r.flatten())
        uniq_vals_g = np.unique(labeled_array_g.flatten())

        # intensity fetch per bacteria
        intens_g = np.minimum(tmp_im_g, base_im_g)
        intens_r = np.minimum(tmp_im_r, base_im_r)

        #print(uniq_vals_r, uniq_vals_g)
        # intensities in the red ch
        for val in uniq_vals_r[uniq_vals_r!=0]:

            idxs = np.argwhere(labeled_array_r == val)
            x,y = idxs[len(idxs) // 2][0], idxs[len(idxs) // 2][1]

            intensity_dict[f"{unique_file}_red_bac_nr_{val}"] = intens_r[x,y]

        # intensities in the green ch
        for val in uniq_vals_g[uniq_vals_g!=0]:

            idxs = np.argwhere(labeled_array_g ==val)
            x,y = idxs[len(idxs) // 2][0], idxs[len(idxs) // 2][1]
      
            
            intensity_dict[f"{unique_file}_green_bac_nr_{val}"] = intens_g[x,y]

print(bac_counting)
#print(intensity_dict)

        # saving thresholded im-s 
        # cv2.imwrite(f"{TARGET_PATH}{unique_file}_th_r.png", tmp_im_r)
        # cv2.imwrite(f"{TARGET_PATH}{unique_file}_th_g.png", tmp_im_g)

        ### saving values to the XLSX file ###
        # current month + year
    #     sheet_name = datetime.now().strftime("%B") +" "+datetime.now().strftime("%Y")

    #     # type bac not possible right now 
    #     nr_of_stacks = len(stacks)
    #     green_red = [bac_counting[f"{unique_file}_green"], bac_counting[f"{unique_file}_red"]]
    #     green_red_ratio =  label(base_im_g)[1] / label(base_im_r)[1]
    #     decision = np.where(green_red_ratio<1, "Dead", "Alive").item()
    #     runtime =  time.time() - start_time

    #     new_row = [f, nr_of_stacks, str(green_red), decision,runtime]

    #     wb = load_workbook(filename)

    #     if sheet_name in wb.sheetnames:

    #         ws = wb.worksheets[-1] # select last worksheet
    #     else:
    #         wb.create_sheet(sheet_name)
    #         headers_row =  ["File path", "Nr of stacks", "Amount of green, red bacteria", "Dead/alive", "Runtime (s)"]
    #         ws = wb.worksheets[-1]
    #         ws.append(headers_row)

    #     ws.append(new_row)
    #     wb.save(filename)
    #     wb.close()

    # start_time = time.time()

    ### moving the CZI file to the processed folder

    # TODO 