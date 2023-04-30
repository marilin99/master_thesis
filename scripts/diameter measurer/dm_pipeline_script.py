import cv2
import numpy as np 
import time 
import os
import re

## helper functions ##
from classical_segmenter import classical_segment
from scale_obtain import scale_obtain 
from thinner import thinner, thinner_2k_5k
from point_picker import point_picker
from dm_finder import dm_finder
from unet_pred import net_prediction

## adding data to the xlsx file ##
from openpyxl import load_workbook
from datetime import datetime

## moving file to the Processed SEMs folder on the Network Drive ##
TARGET_PATH = "/home/marilin/Documents/ESP/data/fiber_tests/fiber_test_2/classical_results_2/"

## To be processed SEMs folder on the Network Drive
#ORIG_PATH = "/home/marilin/Documents/ESP/data/fiber_tests/fiber_test_2/original_data_2/"
ORIG_PATH = "/home/marilin/Documents/ESP/data/fiber_tests/fiber_test_1/orig_sub/"

## Excel workbook path ## 
filename = "/home/marilin/Documents/ESP/data/fiber_tests/fibers_demo.xlsx"

#ORIG_PATH = "/run/user/1000/gvfs/smb-share:server=kivi.ut.ee,share=tensonilabor/Georg L/Fibre diameter_Marilin/Co-Axial  SEM PVA-Pleu/"

FILES = os.listdir(ORIG_PATH) 


for f in FILES:
    if f.endswith(".tif") or f.endswith(".png") or f.endswith(".jpg"):

        start_time = time.time()
        PATH_1 = ORIG_PATH+f
        # assuming an extension is 4 characters (including the ".")
        core_name = f[:-4]


        ## CLASSICAL SEGMENTATION ##
        # otsu thresholding for 2k and 5k if specified in path
        if re.search("(_2k_|_5k_|2k)", PATH_1) != None:
            
            segmented_im = cv2.threshold(cv2.imread(PATH_1, 0), 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1]
            dist, thinned = thinner_2k_5k(segmented_im)
        
        else:
            # classical segmentation for other magnifications
            segmented_im = classical_segment(PATH_1)

            dist, thinned = thinner(segmented_im)

        
        ## U-NET segmentation ##

        # this might skip 15k - needs a better way to assess this
        # if re.search("(_2k_|_5k_|2k)", PATH_1) != None:
        #     continue
        # else:
        #     segmented_im = net_prediction(PATH_1)
        #     print("time taken for unet segmentation", time.time() - start_time)
        #     dist, thinned = thinner(segmented_im)
        #     print("time taken for thinning", time.time() - start_time)

        # 100 measurements per image
        pt_s = point_picker(segmented_im, 100)
        
        # val, unit, px amount - from original image
        scales = scale_obtain(PATH_1)
        try:
            if scales[1] == "um":
                nano_per_px = int(scales[0]) * 1000 / int(scales[-1])
            # scale = nm 
            elif scales[1] == "nm": 
                nano_per_px = int(scales[0]) / int(scales[-1])

        except Exception: 
            with open(f"{TARGET_PATH}{core_name}.txt", "w+") as file_ex:
                file_ex.write("Scaling off, check the image")
            continue


        # leaving the lower part (scale included) in for now
        h,w = segmented_im.shape[:2]

        first_dm_s, first_excs, coords  = dm_finder(thinned, dist, segmented_im, pt_s, h,w,nano_per_px) #[:3]

        ### saving values to the XLSX file ###
        sheet_name = datetime.now().strftime("%B") +" "+datetime.now().strftime("%Y")

        new_row = [core_name, str(first_dm_s), np.mean(first_dm_s), np.std(first_dm_s), np.median(first_dm_s), time.time() - start_time]

        wb = load_workbook(filename)

        if sheet_name in wb.sheetnames:
            ws = wb.worksheets[-1] # select last worksheet
        else:
            wb.create_sheet(sheet_name)
            headers_row =  ["File path", "Diameter measures (nm)", "Mean (nm)", "Standard deviation (nm)", "Median (nm)", "Runtime (s)"]
            ws = wb.worksheets[-1]
            ws.append(headers_row)

        ws.append(new_row)
        wb.save(filename)
        wb.close()

        start_time = time.time()

        ## moving the file to the target path ## 

        # TODO


        #####
    

