from openpyxl import load_workbook
from datetime import datetime

## fiber xlsx 
PATH = "/home/marilin/Documents/ESP/data/fiber_tests/"
filename = f"{PATH}fibers_demo.xlsx"
sheet_name = datetime.now().strftime("%B") +" "+datetime.now().strftime("%Y")
sheet_name = "May 2023"
# new data 
new_row = ['1', '2', '3']

wb = load_workbook(filename)

if sheet_name in wb.sheetnames:
    ws = wb.worksheets[-1] # select last worksheet
else:
    wb.create_sheet(sheet_name)
    headers_row =  ["File path", "Diameter measures (nm)", "Runtime"]
    ws = wb.worksheets[-1]
    ws.append(headers_row)

ws.append(new_row)
wb.save(filename)
wb.close()
    

# bacteria xlsx 
# PATH = "/home/marilin/Documents/ESP/data/bacteria_tests/"
# filename = f"{PATH}bacteria_demo.xlsx"
# sheet_name = datetime.now().strftime("%B") +" "+datetime.now().strftime("%Y")
# # new data 
# new_row = ['1', '2', '3']

# wb = load_workbook(filename)

# if sheet_name in wb.sheetnames:
#     ws = wb.worksheets[-1] # select last worksheet
# else:
#     wb.create_sheet(sheet_name)
#     headers_row = ["File path", "Type of bacteria", "Nr of stacks", "Green/Red bacteria ratio", "Dead/alive", "Runtime"]
#     ws = wb.worksheets[-1]
#     ws.append(headers_row)

# ws.append(new_row)
# wb.save(filename)
# wb.close()